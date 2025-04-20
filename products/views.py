# ==============================
# ✅ CLEANED IMPORTS
# ==============================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction, IntegrityError
from django.db.models import Sum, Max, Q, F, ExpressionWrapper, DecimalField
from django.utils.timezone import now, localdate
from django.urls import reverse
from django.template.loader import render_to_string, get_template
from django.core.exceptions import ObjectDoesNotExist

from decimal import Decimal
from datetime import datetime, timedelta
from io import BytesIO
import csv

from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment




from .models import (
    Product, InventoryEntry, SalesEntry, Order, Invoice, InvoiceItem,
    Seller, SellerInventory, LoadingRecord, UnloadingRecord,
    SellerProductDayEntry, DailySellerStockRecord, SellerPayment,
    Customer, CustomerOrder, CustomerOrderItem, SalesRecord, Expense, Revenue
)

from .forms import (
    ProductForm, InventoryEntryForm, SalesEntryForm,
    InvoiceForm, InvoiceItemForm, ExpenseForm, RevenueForm
)

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect

@login_required
def post_login_redirect(request):
    user = request.user

    if hasattr(user, 'seller'):
        role = getattr(user.seller, 'role', None)

        # Ensure seller role exists and is valid
        if role == 'seller':
            return redirect('products:mobile_landing')
        elif role == 'manager':
            return redirect('products:manager_landing')
        else:
            # Unknown role fallback
            return redirect('products:mobile_landing')

    # If no seller object is linked (possibly superuser/admin)
    return redirect('products:metrics_dashboard')

@login_required
def product_form(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm()
    return render(request, 'products/product_form.html', {'form': form})

@login_required
def product_list(request):
    # Get query parameters for filtering
    category = request.GET.get('category')
    supplier = request.GET.get('supplier')

    # Filter products based on selected category and supplier
    products = Product.objects.all()
    if category:
        products = products.filter(category=category)
    if supplier:
        products = products.filter(supplier=supplier)

    # Get distinct categories and suppliers for the filter dropdowns
    categories = Product.objects.values_list('category', flat=True).distinct()
    suppliers = Product.objects.values_list('supplier', flat=True).distinct()

    return render(request, 'products/product_list.html', {
        'products': products,
        'categories': categories,
        'suppliers': suppliers,
        'selected_category': category,
        'selected_supplier': supplier,
    })

@login_required
def product_edit(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/product_edit.html', {'form': form, 'product': product})

@login_required
def product_delete(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        return redirect('products:product_list')
    return render(request, 'products/product_confirm_delete.html', {'product': product})


@login_required
def inventory_entry(request):
    if request.method == 'POST':
        form = InventoryEntryForm(request.POST)
        if form.is_valid():
            entry = form.save()
            if entry.entry_type == 'restock':
                entry.product.quantity += entry.quantity
            elif entry.entry_type == 'return':
                entry.product.quantity += entry.quantity
            entry.product.save()
            return redirect('products:inventory_entry')
    else:
        form = InventoryEntryForm()
    entries = InventoryEntry.objects.all().order_by('-date')
    return render(request, 'inventory/inventory_entry.html', {'form': form, 'entries': entries})

@login_required
def sales_entry(request):
    if request.method == 'POST':
        form = SalesEntryForm(request.POST)
        if form.is_valid():
            sale = form.save()
            sale.product.quantity -= sale.quantity
            sale.product.save()
            return redirect('products:sales_entry')
    else:
        form = SalesEntryForm()
    sales = SalesEntry.objects.all().order_by('-date')
    return render(request, 'inventory/inventory_sales.html', {'form': form, 'sales': sales})

@login_required
def inventory_entries(request):
    entries = InventoryEntry.objects.all().order_by('-date')
    return render(request, 'inventory/inventory_entries.html', {'entries': entries})

@login_required
def edit_inventory_entry(request, entry_id):
    entry = get_object_or_404(InventoryEntry, id=entry_id)
    if request.method == 'POST':
        form = InventoryEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, "Inventory entry updated successfully!")
            return redirect('inventory:inventory_entries')
    else:
        form = InventoryEntryForm(instance=entry)
    return render(request, 'inventory/edit_inventory_entry.html', {'form': form, 'entry': entry})

@login_required
def delete_inventory_entry(request, entry_id):
    entry = get_object_or_404(InventoryEntry, id=entry_id)
    if request.method == 'POST':
        entry.delete()
        messages.success(request, "Inventory entry deleted successfully!")
        return redirect('products:inventory_entries')
    return render(request, 'inventory/delete_inventory_entry.html', {'entry': entry})

@login_required
def sales_entries(request):
    sales = SalesEntry.objects.all().order_by('-date')
    return render(request, 'inventory/sales_entries.html', {'sales': sales})

@login_required
def edit_sales_entry(request, sale_id):
    sale = get_object_or_404(SalesEntry, id=sale_id)
    if request.method == 'POST':
        form = SalesEntryForm(request.POST, instance=sale)
        if form.is_valid():
            form.save()
            messages.success(request, "Sales entry updated successfully!")
            return redirect('products:sales_entries')
    else:
        form = SalesEntryForm(instance=sale)
    return render(request, 'inventory/edit_sales_entry.html', {'form': form, 'sale': sale})

@login_required
def delete_sales_entry(request, sale_id):
    sale = get_object_or_404(SalesEntry, id=sale_id)
    if request.method == 'POST':
        sale.delete()
        messages.success(request, "Sales entry deleted successfully!")
        return redirect('inventory:sales_entries')
    return render(request, 'inventory/delete_sales_entry.html', {'sale': sale})

@login_required
def inventory_status(request):
    # Get filter parameters
    category_filter = request.GET.get('category')
    supplier_filter = request.GET.get('supplier')

    # Filter products
    products = Product.objects.all()
    if category_filter:
        products = products.filter(category=category_filter)
    if supplier_filter:
        products = products.filter(supplier=supplier_filter)

    # Group products by category
    categories = Product.CATEGORY_CHOICES
    products_by_category = {
        category[0]: products.filter(category=category[0]) for category in categories
    }

    # Get distinct categories and suppliers for filters
    distinct_categories = Product.objects.values_list('category', flat=True).distinct()
    distinct_suppliers = Product.objects.values_list('supplier', flat=True).distinct()

    return render(request, 'inventory/inventory_status.html', {
        'products_by_category': products_by_category,
        'distinct_categories': distinct_categories,
        'distinct_suppliers': distinct_suppliers,
        'selected_category': category_filter,
        'selected_supplier': supplier_filter,
    })

@login_required
def dashboard(request):
    total_products = Product.objects.count()
    total_stock_value = Product.objects.aggregate(
        total_value=Sum('quantity', field="quantity * selling_price")
    )['total_value'] or 0
    total_sales_value = SalesEntry.objects.aggregate(
        total_sales=Sum('quantity', field="quantity * product__selling_price")
    )['total_sales'] or 0
    low_stock_products = Product.objects.filter(quantity__lt=5).count()

    recent_entries = InventoryEntry.objects.order_by('-date')[:5]
    recent_sales = SalesEntry.objects.order_by('-date')[:5]

    context = {
        'total_products': total_products,
        'total_stock_value': total_stock_value,
        'total_sales_value': total_sales_value,
        'low_stock_products': low_stock_products,
        'recent_entries': recent_entries,
        'recent_sales': recent_sales,
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required
def bulk_operations(request):
    if request.method == 'POST':
        operation = request.POST.get('operation')  # Get selected operation (sales or entry)
        quantities = {int(k.split('_')[1]): int(v) for k, v in request.POST.items() if k.startswith('quantity_') and v}

        with transaction.atomic():  # Ensure atomicity of bulk operation
            for product_id, quantity in quantities.items():
                product = Product.objects.get(id=product_id)
                if operation == 'entry':
                    product.quantity += quantity
                    InventoryEntry.objects.create(product=product, entry_type='restock', quantity=quantity)
                elif operation == 'sales':
                    product.quantity -= quantity
                    SalesEntry.objects.create(product=product, quantity=quantity)
                product.save()
        # Add success message
        messages.success(request, f"Bulk {operation.capitalize()} operation completed successfully!")

        return redirect('products:bulk_operations')  # Redirect to the same page after processing

    # Fetch products grouped by category
    categories = Product.objects.values_list('category', flat=True).distinct()
    products_by_category = {category: Product.objects.filter(category=category) for category in categories}

    return render(request, 'inventory/bulk_operations.html', {
        'products_by_category': products_by_category
    })

@login_required
def generate_order_number():
    current_year = now().year % 100  # Get the last two digits of the current year, e.g., 2025 -> 25

    # Get the last order number for the current year
    last_order = Order.objects.filter(order_number__endswith=f"/{current_year}").aggregate(
        max_number=Max("order_number")
    )
    last_number = last_order["max_number"]

    if last_number:
        # Extract the XXXX part and increment it
        last_number_int = int(last_number.split("/")[0][2:])  # Extract XXXX from BCXXXX/YY
        next_number = last_number_int + 1
    else:
        # No orders for the current year, start from 0001
        next_number = 1

    # Format the order number as BCXXXX/YY
    return f"BC{next_number:04d}/{current_year}"

@login_required
def reorder_page(request):
    if request.method == 'POST':
        # Get supplier and order data from the request
        supplier = request.POST.get('supplier', 'Unknown Supplier')
        products = Product.objects.all()  # Retrieve all products

        # Generate a unique order number
        order_number = generate_order_number()  # Reuse the generate_order_number function

        # Prepare products with non-zero quantities
        selected_products = [
            {
                'name': product.name,
                'purchase_price': product.purchase_price,
                'quantity': int(request.POST.get(f'quantity_{product.id}', '0') or '0'),
                'total': float(product.purchase_price) * int(request.POST.get(f'quantity_{product.id}', '0') or '0'),
            }
            for product in products
            if int(request.POST.get(f'quantity_{product.id}', '0') or '0') > 0
        ]

        # Calculate total amount
        total_amount = sum(product['total'] for product in selected_products)

        # Save the order to the database
        order = Order.objects.create(
            order_number=order_number,
            supplier=supplier,
            total_amount=total_amount,
        )

        # Generate PDF for the order
        context = {
            'order': order,
            'products': selected_products,
            'total_amount': total_amount,
        }

        response = generate_pdf('products/reorder_pdf.html', context)

        return response

    return render(request, 'products/reorder_page.html', {'products': Product.objects.all()})

@login_required
def generate_pdf(template_path, context):
    template = render_to_string(template_path, context)
    pdf_file = BytesIO()
    pisa.CreatePDF(BytesIO(template.encode('UTF-8')), pdf_file)
    pdf_file.seek(0)
    return HttpResponse(pdf_file, content_type='application/pdf')

@login_required
def order_list(request):
    query = request.GET.get('search', '')  # Get the search query from the request
    if query:
        orders = Order.objects.filter(Q(order_number__icontains=query))
    else:
        orders = Order.objects.all().order_by('-created_at')

    return render(request, 'products/order_list.html', {'orders': orders, 'search_query': query})

@login_required
def order_details(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'products/order_details.html', {'order': order})

@login_required
def order_edit(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    if request.method == 'POST':
        order.supplier = request.POST.get('supplier', order.supplier)
        order.status = request.POST.get('status', order.status)
        order.save()
        return redirect('products:order_list')

    return render(request, 'products/order_edit.html', {'order': order})

@login_required
def order_delete(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    if request.method == 'POST':
        order.delete()
        return redirect('products:order_list')

    return render(request, 'products/order_confirm_delete.html', {'order': order})

@login_required
def import_products_from_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "The file uploaded is not a CSV file.")
            return render(request, 'products/import_products.html')

        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            # import pdb; pdb.set_trace()
            for row in reader:
                try:
                    Product.objects.create(
                        code=row['\ufeffcode'],
                        name=row['name'],
                        name_ar=row.get('name_ar', None),
                        purchase_price=float(row['purchase_price']),
                        selling_price=float(row['selling_price']),
                        quantity=int(row['quantity']),
                        category=row['category'],
                        supplier=row['supplier']
                    )
                except Exception as e:
                    messages.error(request, f"Error importing product {row.get('name', 'Unknown')}: {str(e)}")
            messages.success(request, "Products imported successfully!")
        except Exception as e:
            messages.error(request, f"Error processing the file: {str(e)}")
        
        return render(request, 'products/import_products.html')
    
    # Render the upload page for GET requests
    return render(request, 'products/import_products.html')

@login_required
def export_inventory_to_pdf(request):
    category = request.GET.get('category')
    supplier = request.GET.get('supplier')

    # Filter products based on selected category and supplier
    products = Product.objects.all()
    if category:
        products = products.filter(category=category)
    if supplier:
        products = products.filter(supplier=supplier)

    # Group products by category
    categories = Product.objects.values_list('category', flat=True).distinct()
    products_by_category = {cat: products.filter(category=cat) for cat in categories}

    context = {
        'products_by_category': products_by_category,
    }

    # Render PDF
    template_path = 'inventory/inventory_pdf.html'
    html = render_to_string(template_path, context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="inventory.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', content_type='text/plain')
    return response

@login_required
def invoice_list(request):
    invoices = Invoice.objects.all().order_by('-date')
    return render(request, 'products/invoice_list.html', {'invoices': invoices})

@login_required
def new_invoice(request):
    if request.method == 'POST':
        client = request.POST.get('client')
        date = request.POST.get('date')
        details = request.POST.get('details')

        # Create the invoice
        invoice = Invoice.objects.create(client=client, date=date, details=details)

        # Add items to the invoice
        for product_id, quantity in request.POST.items():
            if product_id.startswith('quantity_') and quantity:
                product_id = int(product_id.split('_')[1])
                product = Product.objects.get(id=product_id)
                unit_price = product.selling_price
                quantity = int(quantity)
                InvoiceItem.objects.create(
                    invoice=invoice,
                    product=product,
                    unit_price=unit_price,
                    quantity=quantity,
                )

        # Update total amount
        invoice.total_amount = sum(item.total_price for item in invoice.items.all())
        invoice.save()

        return redirect('products:invoice_list')

    # Fetch products grouped by category
    products = Product.objects.all()
    categories = Product.CATEGORY_CHOICES
    products_by_category = {cat[0]: products.filter(category=cat[0]) for cat in categories}

    return render(request, 'products/new_invoice.html', {'products_by_category': products_by_category})

@login_required
def invoice_details(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = InvoiceItem.objects.filter(invoice=invoice)

    return render(request, 'products/invoice_details.html', {
        'invoice': invoice,
        'items': items
    })

@login_required
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    if request.method == 'POST':
        invoice_form = InvoiceForm(request.POST, instance=invoice)

        if invoice_form.is_valid():
            invoice_form.save()

            # Update invoice items
            items = InvoiceItem.objects.filter(invoice=invoice)
            for item in items:
                item.quantity = int(request.POST.get(f'quantity_{item.id}', item.quantity))
                item.save()

            messages.success(request, "Invoice updated successfully!")
            return redirect('products:invoice_list')

    else:
        invoice_form = InvoiceForm(instance=invoice)
        items = InvoiceItem.objects.filter(invoice=invoice)

    return render(request, 'products/edit_invoice.html', {
        'invoice_form': invoice_form,
        'invoice': invoice,
        'items': items
    })

@login_required
def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    if request.method == 'POST':
        invoice.delete()
        messages.success(request, "Invoice deleted successfully!")
        return redirect('products:invoice_list')

    return render(request, 'products/delete_invoice.html', {
        'invoice': invoice
    })

@login_required
def load_inventory(request):
    if request.method == 'POST':
        seller_id = request.POST.get('seller')
        seller = get_object_or_404(Seller, id=seller_id)
        
        for product in Product.objects.all():
            quantity = int(request.POST.get(f'quantity_{product.id}', 0))
            if quantity > 0:
                # Update warehouse inventory
                product.quantity -= quantity
                product.save()
                
                # Add to seller's inventory
                seller_inventory, created = SellerInventory.objects.get_or_create(
                    seller=seller, product=product
                )
                seller_inventory.quantity += quantity
                seller_inventory.save()

                # Create loading record
                LoadingRecord.objects.create(seller=seller, product=product, quantity=quantity)

        messages.success(request, "Stock chargé avec succès!")
        return redirect('products:load_inventory')
    
    sellers = Seller.objects.all()
    products = Product.objects.all()
    return render(request, 'inventory/load_inventory.html', {'sellers': sellers, 'products': products})

@login_required
def unload_inventory(request):
    seller_id = request.GET.get('seller') or request.POST.get('seller')  # Ensure we get seller from GET or POST
    seller_inventory = []

    if seller_id:
        try:
            seller = Seller.objects.get(id=seller_id)
        except Seller.DoesNotExist:
            messages.error(request, "Vendeur invalide sélectionné.")
            return redirect('products:unload_inventory')
        
        seller_inventory = SellerInventory.objects.filter(seller=seller)

    if request.method == 'POST':
        seller_id = request.POST.get('seller')

        if not seller_id:
            messages.error(request, "Veuillez sélectionner un vendeur avant de soumettre.")
            return redirect('products:unload_inventory')

        seller = get_object_or_404(Seller, id=seller_id)
        sales_records = []

        for item in SellerInventory.objects.filter(seller=seller):
            loaded_quantity = item.quantity  # Quantity loaded
            unloaded_quantity = int(request.POST.get(f'quantity_{item.product.id}', 0) or 0)  # Default to 0 if empty

            if unloaded_quantity > 0:
                item.quantity -= unloaded_quantity
                item.save()

                item.product.quantity += unloaded_quantity  # Return to warehouse
                item.product.save()

                UnloadingRecord.objects.create(
                    seller=seller,
                    product=item.product,
                    quantity=unloaded_quantity,
                    date=now().date()
                )

                # Calculate quantity sold
                quantity_sold = loaded_quantity - unloaded_quantity
                if quantity_sold > 0:
                    sales_records.append(SalesRecord(
                        seller=seller,
                        product=item.product,
                        quantity_sold=quantity_sold,
                        total_amount=quantity_sold * item.product.selling_price,
                        date=now().date()
                    ))

        # Save all sales records in bulk
        SalesRecord.objects.bulk_create(sales_records)

        messages.success(request, "Stock déchargé et ventes calculées avec succès!")
        return redirect('products:unload_inventory')

    sellers = Seller.objects.all()
    return render(request, 'inventory/unload_inventory.html', {
        'sellers': sellers,
        'seller_inventory': seller_inventory,
        'selected_seller_id': seller_id  # Ensure seller remains selected
    })

@login_required
def sales_report(request):
    selected_date = request.GET.get('date', now().date())

    sales = SalesEntry.objects.filter(date=selected_date).values(
        'date', 'product__name'
    ).annotate(
        total_quantity_sold=Sum('quantity'),
        total_sales=Sum('quantity') * Sum('product__selling_price')
    ).order_by('date')

    return render(request, 'inventory/sales_report.html', {
        'sales': sales,
        'selected_date': selected_date
    })

@login_required
def sale_details(request, sale_id):
    sale = get_object_or_404(SalesEntry, id=sale_id)
    total_amount = sale.quantity * sale.product.selling_price  # Calculate total

    return render(request, 'sales/sale_details.html', {
        'sale': sale,
        'total_amount': total_amount
    })

@login_required
def manager_inventory(request):
    seller_id = request.GET.get("seller")
    selected_seller = None
    unique_dates = []

    if seller_id:
        selected_seller = get_object_or_404(Seller, id=seller_id)

        # Fetch dates directly without TruncDate()
        load_dates = LoadingRecord.objects.filter(seller=selected_seller).values_list("date", flat=True).distinct()
        unload_dates = UnloadingRecord.objects.filter(seller=selected_seller).values_list("date", flat=True).distinct()

        # Merge and sort dates
        unique_dates = sorted(set(load_dates).union(set(unload_dates)), reverse=True)

    sellers = Seller.objects.all()
    return render(request, "inventory/manager_inventory.html", {
        "sellers": sellers,
        "selected_seller": selected_seller,
        "unique_dates": unique_dates,
    })

@login_required
def load_new_inventory(request):
    sellers = Seller.objects.all()
    products = Product.objects.all()
    voiture_quantities = {}

    selected_seller_id = request.GET.get('seller') or request.POST.get('seller')
    selected_date_str = request.GET.get('date') or request.POST.get('date')
    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date() if selected_date_str else now().date()

    if selected_seller_id:
        selected_seller = get_object_or_404(Seller, id=selected_seller_id)

        yesterday = selected_date - timedelta(days=1)
        for product in products:
            yesterday_entry = SellerProductDayEntry.objects.filter(
                seller=selected_seller, product=product, date=yesterday
            ).first()
            voiture_quantities[product.id] = yesterday_entry.retour if yesterday_entry else 0

        if request.method == 'POST':
            entries = []
            for product in products:
                sortie_qty = int(request.POST.get(f'sortie_{product.id}', 0))
                voiture_qty = voiture_quantities.get(product.id, 0)

                entry, _ = SellerProductDayEntry.objects.update_or_create(
                    seller=selected_seller,
                    product=product,
                    date=selected_date,
                    defaults={
                        'voiture': voiture_qty,
                        'sortie': sortie_qty,
                        'retour': 0,
                    }
                )
                entry.total_loaded = voiture_qty + sortie_qty
                entries.append(entry)
            filtered_entries = [entry for entry in entries if entry.total_loaded > 0]

            # Generate PDF
            template = get_template('pdf/load_report.html')
            html = template.render({
                'seller': selected_seller,
                'date': selected_date,
                'entries': filtered_entries,
            })

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'filename="Chargement-{selected_seller.name}-{selected_date}.pdf"'
            pisa.CreatePDF(html, dest=response)
            return response

    return render(request, 'inventory/load_new_inventory.html', {
        'sellers': sellers,
        'products': products,
        'voiture_quantities': voiture_quantities,
        'selected_seller_id': selected_seller_id,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
    })

@login_required
def unload_new_inventory(request):
    seller_id = request.GET.get("seller")
    date_str = request.GET.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else now().date()

    seller = None
    voiture_quantities = {}
    sortie_quantities = {}

    if seller_id:
        seller = get_object_or_404(Seller, id=seller_id)

        # Fetch voiture = yesterday's retour
        yesterday = selected_date - timedelta(days=1)
        previous_unloads = SellerProductDayEntry.objects.filter(seller=seller, date=yesterday)
        for entry in previous_unloads:
            voiture_quantities[entry.product.id] = entry.retour

        # Fetch sortie = today's sortie
        today_loads = SellerProductDayEntry.objects.filter(seller=seller, date=selected_date)
        for entry in today_loads:
            sortie_quantities[entry.product.id] = entry.sortie

        if request.method == "POST":
            for product in Product.objects.all():
                retour_qty = int(request.POST.get(f"retour_{product.id}", 0))

                entry, _ = SellerProductDayEntry.objects.get_or_create(
                    seller=seller, product=product, date=selected_date
                )
                entry.retour = retour_qty
                entry.save()

            messages.success(request, "Déchargement enregistré avec succès.")

            # Redirect directly to PDF
            pdf_url = reverse('products:export_unload_pdf', args=[seller.id, selected_date])
            return redirect(pdf_url)

    return render(request, "inventory/unload_new_inventory.html", {
        "sellers": Seller.objects.all(),
        "selected_seller": seller,
        "products": Product.objects.all(),
        "selected_date": selected_date,
        "voiture": voiture_quantities,
        "sortie": sortie_quantities,
    })

@login_required
def daily_seller_stock(request):
    # Parse date (default today)
    date_str = request.GET.get('date')
    seller_id = request.GET.get('seller')
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else now().date()

    sellers = Seller.objects.all()
    products = Product.objects.all()
    selected_seller = Seller.objects.filter(id=seller_id).first() if seller_id else None
    previous_date = selected_date - timedelta(days=1)

    records = []
    if request.method == 'POST' and selected_seller:
        for product in products:
            sortie = int(request.POST.get(f'sortie_{product.id}', 0))
            retour = int(request.POST.get(f'retour_{product.id}', 0))

            # Get voiture value from form if user modified it manually
            voiture = int(request.POST.get(f'voiture_{product.id}', 0))

            record, created = DailySellerStockRecord.objects.get_or_create(
                seller=selected_seller,
                product=product,
                date=selected_date,
                defaults={
                    'voiture': voiture,
                    'sortie': sortie,
                    'retour': retour,
                }
            )
            if not created:
                record.voiture = voiture
                record.sortie = sortie
                record.retour = retour
                record.save()

        messages.success(request, "Stock du jour mis à jour avec succès.")
        return redirect(f"?seller={selected_seller.id}&date={selected_date}")

    if selected_seller:
        for product in products:
            # Get today's or create empty
            record = DailySellerStockRecord.objects.filter(
                seller=selected_seller, product=product, date=selected_date
            ).first()

            if not record:
                voiture = DailySellerStockRecord.objects.filter(
                    seller=selected_seller, product=product, date=previous_date
                ).values_list('retour', flat=True).first() or 0
                record = DailySellerStockRecord(
                    seller=selected_seller, product=product, date=selected_date,
                    voiture=voiture, sortie=0, retour=0
                )
            records.append(record)

    return render(request, 'inventory/daily_seller_stock.html', {
        'sellers': sellers,
        'products': products,
        'records': records,
        'selected_seller': selected_seller,
        'selected_date': selected_date
    })

@login_required
def daily_sales_summary(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    filters = {}
    if start_date:
        filters['date__gte'] = start_date
    if end_date:
        filters['date__lte'] = end_date

    summary = (
        SellerProductDayEntry.objects
        .filter(**filters)
        .values('date', 'seller__name', 'seller_id')
        .annotate(
            total_vendu=Sum('vendu'),
            total_amount=Sum('amount')
        )
        .order_by('-date', 'seller__name')
    )

    return render(request, 'inventory/daily_sales_summary.html', {
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date
    })

@login_required
def daily_sales_detail(request, seller_id, date):
    seller = get_object_or_404(Seller, id=seller_id)
    entries = SellerProductDayEntry.objects.filter(seller=seller, date=date).select_related('product')

    return render(request, 'inventory/daily_sales_detail.html', {
        'seller': seller,
        'date': date,
        'entries': entries,
    })

@login_required
def export_unload_pdf(request, seller_id, date):
    date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    seller = get_object_or_404(Seller, id=seller_id)

    all_entries = SellerProductDayEntry.objects.filter(
        seller=seller, date=date_obj
    ).select_related('product')

    # Filter only entries where vendu > 0
    filtered_entries = []
    total_amount = 0
    for entry in all_entries:
        entry.vendu = entry.voiture + entry.sortie - entry.retour
        entry.amount = entry.vendu * entry.product.selling_price
        if entry.vendu > 0:
            filtered_entries.append(entry)
            total_amount += entry.amount

    context = {
        'seller': seller,
        'date': date_obj,
        'entries': filtered_entries,
        'total_amount': total_amount
    }

    template_path = 'pdf/unload_report.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{seller.name}-{date}.pdf"'

    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse(f'Erreur PDF: {pisa_status.err}')
    return response


@login_required
def metrics_dashboard(request):
    today = now().date()

    total_products = Product.objects.count()
    low_stock_products = Product.objects.filter(quantity__lte=10)
    low_stock_count = low_stock_products.count()

    today_entries = SellerProductDayEntry.objects.filter(date=today)

    total_retour = today_entries.aggregate(r=Sum('retour'))['r'] or 0
    total_loaded = today_entries.aggregate(
        loaded=Sum(F('voiture') + F('sortie'))
    )['loaded'] or 0

    retour_rate = (total_retour / total_loaded * 100) if total_loaded > 0 else 0

    active_sellers = Seller.objects.filter(sellerproductdayentry__date=today).distinct()
    active_sellers_count = active_sellers.count()

    top_products = (
        today_entries.values('product__name')
        .annotate(total_vendu=Sum('vendu'))
        .order_by('-total_vendu')[:5]
    )

    sales_by_seller = (
        today_entries.values('seller__name')
        .annotate(total_vendu=Sum('vendu'), total_amount=Sum('amount'))
        .order_by('-total_amount')
    )

    current_month = today.month
    current_year = today.year
    past_days_entries = SellerProductDayEntry.objects.filter(
        date__lt=today,
        date__year=current_year,
        date__month=current_month
    )

    monthly_sales_amount = past_days_entries.aggregate(amount=Sum('amount'))['amount'] or 0

    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'low_stock_count': low_stock_count,
        'active_sellers_count': active_sellers_count,
        'top_products': top_products,
        'seller_sales_today': sales_by_seller,
        'today': today,
        'monthly_sales_amount': monthly_sales_amount,
        'retour_rate': retour_rate,
    }
    return render(request, 'inventory/metrics_dashboard.html', context)


@login_required
def seller_payment_entry(request):
    seller_id = request.GET.get("seller")
    date_str = request.GET.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else now().date()

    seller = get_object_or_404(Seller, id=seller_id) if seller_id else None
    expected_amount = 0
    existing_payment = None

    if seller:
        expected_amount = SellerProductDayEntry.objects.filter(
            seller=seller, date=selected_date
        ).aggregate(total=Sum('amount'))['total'] or 0

        existing_payment = SellerPayment.objects.filter(
            seller=seller, date=selected_date
        ).first()

        if request.method == "POST":
            paid_str = request.POST.get("paid_amount")
            paid_amount = Decimal(paid_str) if paid_str else 0

            SellerPayment.objects.update_or_create(
                seller=seller,
                date=selected_date,
                defaults={
                    'paid_amount': paid_amount,
                    'expected_amount': expected_amount
                }
            )

            messages.success(request, "Paiement enregistré.")
            return redirect(f"{reverse('products:seller_payment_history')}?seller={seller.id}")

    return render(request, "payments/payment_entry.html", {
        "sellers": Seller.objects.all(),
        "selected_seller": seller,
        "selected_date": selected_date,
        "expected_amount": expected_amount,
        "existing_payment": existing_payment,
    })

@login_required
def unpaid_balances_report(request):
    unpaid_summary = (
        SellerPayment.objects
        .values('seller__name', 'seller__id')
        .annotate(
            total_expected=Sum('expected_amount'),
            total_paid=Sum('paid_amount'),
            total_remaining=ExpressionWrapper(
                Sum(F('expected_amount')) - Sum(F('paid_amount')),
                output_field=DecimalField()
            ),
            last_payment_date=Max('date')
        )
        .filter(
            expected_amount__gt=F('paid_amount')
        )
        .order_by('-total_remaining')
    )

    return render(request, 'payments/unpaid_balances_report.html', {
        'unpaid_summary': unpaid_summary
    })

@login_required
def seller_payment_history(request):
    seller_id = request.GET.get('seller')
    seller = get_object_or_404(Seller, id=seller_id) if seller_id else None

    sold_by_day = {}
    paid_by_day = {}

    if seller:
        sold = (
            SellerProductDayEntry.objects
            .filter(seller=seller)
            .values('date')
            .annotate(total_sold=Sum('amount'))
        )
        for row in sold:
            sold_by_day[row['date']] = row['total_sold']

        payments = (
            SellerPayment.objects
            .filter(seller=seller)
            .values('date')
            .annotate(total_paid=Sum('paid_amount'))
        )
        for row in payments:
            paid_by_day[row['date']] = row['total_paid']

    # Combine all unique dates
    all_dates = sorted(set(sold_by_day.keys()) | set(paid_by_day.keys()))

    history = []
    for date in all_dates:
        sold = sold_by_day.get(date, 0)
        paid = paid_by_day.get(date, 0)
        history.append({
            'date': date,
            'sold': sold,
            'paid': paid,
            'remaining': sold - paid
        })

    context = {
        'sellers': Seller.objects.all(),
        'selected_seller': seller,
        'history': history
    }
    return render(request, 'payments/seller_payment_history.html', context)

@login_required
def customer_list(request):
    try:
        seller = request.user.seller
        customers = Customer.objects.filter(seller=seller)
    except (AttributeError, ObjectDoesNotExist):
        customers = Customer.objects.none()  # Or show all if admin
    
    return render(request, "sales/customer_list.html", {
        "customers": customers
    })

@login_required
def customer_create(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        address = request.POST.get("address")

        try:
            seller = request.user.seller
        except ObjectDoesNotExist:
            messages.error(request, "Aucun vendeur n'est lié à ce compte utilisateur.")
            return redirect("products:customer_list")

        customer = Customer.objects.create(name=name, phone=phone, address=address, seller=seller)
        print(f"✅ Nouveau client ajouté : {customer.name}")
        messages.success(request, "Client ajouté avec succès.")
        return redirect("products:customer_list")

    return render(request, "sales/customer_form.html")

@login_required
def order_create(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id, seller=request.user.seller)
    products = Product.objects.all()

    if request.method == "POST":
        total = Decimal("0.00")
        order = CustomerOrder.objects.create(
            seller=request.user.seller,
            customer=customer,
            status="completed",  # or use a form/select for status
            total_amount=0  # temporary
        )

        for product in products:
            qty_str = request.POST.get(f"product_{product.id}")
            if qty_str:
                quantity = int(qty_str)
                if quantity > 0:
                    total_price = product.selling_price * quantity
                    CustomerOrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        unit_price=product.selling_price,
                        total_price=total_price
                    )
                    total += total_price

        order.total_amount = total
        order.save()

        messages.success(request, "Commande enregistrée avec succès.")
        return redirect("products:order_detail", order_id=order.id)

    return render(request, "sales/order_create.html", {
        "customer": customer,
        "products": products
    })

@login_required
def order_detail(request, order_id):
    order = get_object_or_404(CustomerOrder, id=order_id)
    return render(request, "sales/order_detail.html", {"order": order})

@login_required
def order_pdf(request, order_id):
    order = get_object_or_404(CustomerOrder.objects.select_related('customer', 'seller').prefetch_related('items__product'), id=order_id)

    template = get_template("pdf/order_receipt.html")
    html = template.render({"order": order})

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="Order-{order.id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF", status=500)
    return response

# views.py
@login_required
def order_receipt(request, order_id):
    order = get_object_or_404(CustomerOrder, id=order_id, seller=request.user.seller)
    return render(request, 'pdf/order_receipt_mini.html', {'order': order})

@login_required
def customer_order_list(request):
    sellers = Seller.objects.all()
    customers = Customer.objects.all()

    seller_id = request.GET.get('seller')
    customer_id = request.GET.get('customer')
    date = request.GET.get('date')

    orders = CustomerOrder.objects.all().select_related('customer', 'seller')

    if seller_id:
        orders = orders.filter(seller_id=seller_id)
    if customer_id:
        orders = orders.filter(customer_id=customer_id)
    if date:
        orders = orders.filter(date=date)

    return render(request, 'sales/customer_order_list.html', {
        'orders': orders,
        'sellers': sellers,
        'customers': customers,
        'selected_seller': int(seller_id) if seller_id else None,
        'selected_customer': int(customer_id) if customer_id else None,
        'selected_date': date,
    })

@login_required
def export_load_pdf(request, seller_id, date):
    date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    seller = get_object_or_404(Seller, id=seller_id)

    entries = SellerProductDayEntry.objects.filter(seller=seller, date=date_obj).select_related('product')

    for entry in entries:
        entry.total_loaded = entry.voiture + entry.sortie

    context = {
        'seller': seller,
        'date': date_obj,
        'entries': entries,
    }

    template = get_template('pdf/load_report.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="Chargement-{seller.name}-{date}.pdf"'
   
@login_required
def mobile_load_inventory(request):
    seller = get_object_or_404(Seller, user=request.user)
    products = Product.objects.all()

    if request.method == "POST":
        for product in products:
            qty = int(request.POST.get(f"quantity_{product.id}", 0))
            SellerInventory.objects.update_or_create(
                seller=seller,
                product=product,
                defaults={"quantity": qty}
            )
        messages.success(request, "Inventaire initial chargé.")
        return redirect("products:mobile_inventory_status")

    current_stock = {
        inv.product.id: inv.quantity
        for inv in SellerInventory.objects.filter(seller=seller)
    }

    return render(request, "mobile/load_inventory.html", {
        "products": products,
        "current_stock": current_stock,
    })

@login_required
def mobile_unload_inventory(request):
    seller = get_object_or_404(Seller, user=request.user)
    products = Product.objects.all()

    if request.method == "POST":
        for product in products:
            unload_qty = int(request.POST.get(f"unload_{product.id}", 0))
            inventory = SellerInventory.objects.filter(seller=seller, product=product).first()
            if inventory:
                inventory.quantity = max(inventory.quantity - unload_qty, 0)
                inventory.save()
        messages.success(request, "Déchargement effectué avec succès.")
        return redirect("products:mobile_inventory_status")

    inventory_data = SellerInventory.objects.filter(seller=seller).select_related("product")
    return render(request, "mobile/unload_inventory.html", {
        "inventory": inventory_data
    })

@login_required
def mobile_inventory_status(request):
    seller = get_object_or_404(Seller, user=request.user)
    inventory_data = SellerInventory.objects.filter(seller=seller).select_related("product")
    return render(request, "mobile/inventory_status.html", {
        "inventory": inventory_data
    })

@login_required
def mobile_clients(request):
    seller = get_object_or_404(Seller, user=request.user)
    clients = Customer.objects.filter(seller=seller)

    query = request.GET.get('q')
    if query:
        clients = clients.filter(name__icontains=query)

    return render(request, "mobile/mobile_clients.html", {
        "clients": clients
    })

@login_required
def mobile_orders(request):
    seller = get_object_or_404(Seller, user=request.user)
    orders = CustomerOrder.objects.filter(seller=seller).select_related("customer")

    customer_id = request.GET.get("customer_id")
    date = request.GET.get("date")

    if customer_id:
        orders = orders.filter(customer_id=customer_id)
    if date:
        orders = orders.filter(date__date=date)

    orders = orders.order_by("-date")
    customers = Customer.objects.filter(seller=seller)

    return render(request, "mobile/mobile_orders.html", {
        "orders": orders,
        "customers": customers,
        "selected_customer": customer_id,
        "selected_date": date,
    })

@login_required
def mobile_cash(request):
    seller = get_object_or_404(Seller, user=request.user)
    payments = SellerPayment.objects.filter(seller=seller).order_by("-date")
    total_expected = payments.aggregate(Sum("expected_amount"))['expected_amount__sum'] or 0
    total_paid = payments.aggregate(Sum("paid_amount"))['paid_amount__sum'] or 0
    total_balance = total_expected - total_paid

    return render(request, "mobile/mobile_cash.html", {
        "payments": payments,
        "total_expected": total_expected,
        "total_paid": total_paid,
        "total_balance": total_balance
    })

@login_required
def mobile_create_order(request):
    seller = request.user.seller
    customers = Customer.objects.filter(seller=seller)
    products = Product.objects.all()

    # Capture selected customer from GET query
    selected_customer_id = request.GET.get("customer_id")

    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        customer = get_object_or_404(Customer, id=customer_id, seller=seller)

        order = CustomerOrder.objects.create(
            seller=seller,
            customer=customer,
            total_amount=0,
        )

        total = Decimal('0.00')

        for product in products:
            qty_str = request.POST.get(f'quantity_{product.id}')
            if qty_str:
                qty = int(qty_str)
                if qty > 0:
                    item = CustomerOrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=qty,
                        unit_price=product.selling_price,
                        total_price=product.selling_price * qty
                    )
                    total += item.total_price

        order.total_amount = total
        order.save()

        return redirect('products:mobile_order_detail', order_id=order.id)

    return render(request, 'mobile/create_order.html', {
        'customers': customers,
        'products': products,
        'selected_customer_id': selected_customer_id,
    })

@login_required
def mobile_order_detail(request, order_id):
    order = get_object_or_404(CustomerOrder, id=order_id, seller=request.user.seller)
    return render(request, 'mobile/order_detail.html', {
        'order': order,
        'items': order.items.select_related('product')
    })

@login_required
def mobile_landing(request):
    return render(request, 'mobile/mobile_landing.html')

@login_required
def mobile_create_customer(request):
    seller = request.user.seller
    customer_id = request.GET.get("customer_id")
    is_edit = bool(customer_id)

    customer = get_object_or_404(Customer, id=customer_id, seller=seller) if is_edit else None

    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')

        if is_edit and customer:
            customer.name = name
            customer.phone = phone
            customer.address = address
            customer.save()
            messages.success(request, "Client mis à jour avec succès.")
        else:
            Customer.objects.create(name=name, phone=phone, address=address, seller=seller)
            messages.success(request, "Client ajouté avec succès.")

        return redirect('products:mobile_clients')

    return render(request, 'mobile/mobile_clients_create.html', {
        'customer': customer,
        'is_edit': is_edit,
    })

@login_required
def manager_landing(request):
    today = now().date()
    current_month = today.month
    current_year = today.year

    # Stock value = quantity * selling_price
    stock_value = Product.objects.aggregate(
        total=Sum(F('quantity') * F('selling_price'), output_field=DecimalField())
    )['total'] or 0

    # Monthly sales = sum of amount for current month
    monthly_entries = SellerProductDayEntry.objects.filter(date__year=current_year, date__month=current_month)
    monthly_sales = monthly_entries.aggregate(total=Sum('amount'))['total'] or 0

    # Monthly profit = (selling_price - purchase_price) * vendu
    profit_expr = ExpressionWrapper(
        (F('product__selling_price') - F('product__purchase_price')) * F('vendu'),
        output_field=DecimalField()
    )
    monthly_profit = monthly_entries.aggregate(profit=Sum(profit_expr))['profit'] or 0

    # Low stock products
    low_stock_count = Product.objects.filter(quantity__lte=F('critical_threshold')).count()

    # Top 10 products this month
    top_products = (
        monthly_entries.values('product__name')
        .annotate(total_sold=Sum('vendu'))
        .order_by('-total_sold')[:10]
    )

    # Sales per seller per date
    sales_by_seller = (
        monthly_entries.values('seller__name', 'date')
        .annotate(total=Sum('amount'))
        .order_by('-date')
    )

    context = {
        'total_stock_value': stock_value,
        'monthly_sales': monthly_sales,
        'monthly_profit': monthly_profit,
        'low_stock_count': low_stock_count,
        'top_products': top_products,
        'sales_by_seller': sales_by_seller,
        'today': today,
    }
    return render(request, 'dashboard/manager_landing.html', context)

@login_required
def expense_entry(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('products:expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'expenses/expense_form.html', {'form': form})

@login_required
def expense_list(request):
    category = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    expenses = Expense.objects.all()

    if category:
        expenses = expenses.filter(category=category)
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)

    return render(request, 'expenses/expense_list.html', {
        'expenses': expenses,
        'categories': Expense.CATEGORY_CHOICES,
        'selected_category': category,
        'start_date': start_date,
        'end_date': end_date
    })

@login_required
def expense_list_pdf(request):
    category = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    expenses = Expense.objects.all()

    if category:
        expenses = expenses.filter(category=category)
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        expenses = expenses.filter(date__lte=end_date)

    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or 0


    context = {
        'depenses': expenses,
        'total_amount': total_amount,
        'start_date': start_date,
        'end_date': end_date
    }

    html = render_to_string("expenses/expense_list_pdf.html", context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="dépenses.pdf"'
    pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)
    return response

@login_required
def export_expenses_excel(request):
    category = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    expenses = Expense.objects.all()
    if category:
        expenses = expenses.filter(category=category)
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dépenses"

    headers = ["Date", "Catégorie", "Description", "Montant (DH)"]
    ws.append(headers)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)

    total = 0
    for expense in expenses:
        ws.append([
            expense.date.strftime("%d/%m/%Y"),
            expense.category,
            expense.description or "-",
            float(expense.amount)
        ])
        total += expense.amount

    # Total row
    ws.append(["", "", "TOTAL", total])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="depenses.xlsx"'
    wb.save(response)
    return response

@login_required
def revenue_create(request):
    form = RevenueForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Recette enregistrée avec succès.")
        return redirect('products:revenue_list')

    return render(request, 'revenue/revenue_form.html', {'form': form})

@login_required
def revenue_list(request):
    category = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    revenues = Revenue.objects.all()
    if category:
        revenues = revenues.filter(category=category)
    if start_date:
        revenues = revenues.filter(date__gte=start_date)
    if end_date:
        revenues = revenues.filter(date__lte=end_date)

    categories = Revenue.CATEGORY_CHOICES

    return render(request, 'revenue/revenue_list.html', {
        'revenues': revenues,
        'categories': categories,
        'selected_category': category,
        'start_date': start_date,
        'end_date': end_date
    })

@login_required
def revenue_list_pdf(request):
    category = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    revenues = Revenue.objects.all()
    if category:
        revenues = revenues.filter(category=category)
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        revenues = revenues.filter(date__gte=start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        revenues = revenues.filter(date__lte=end_date)

    total_amount = revenues.aggregate(total=Sum('amount'))['total'] or 0


    context = {
        'revenues': revenues,
        'total_amount': total_amount,
        'start_date': start_date,
        'end_date': end_date
    }

    html = render_to_string("revenue/revenue_list_pdf.html", context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="revenus.pdf"'
    pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)
    return response

@login_required
def export_revenues_excel(request):
    category = request.GET.get("category")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    revenues = Revenue.objects.all()
    if category:
        revenues = revenues.filter(category=category)
    if start_date:
        revenues = revenues.filter(date__gte=start_date)
    if end_date:
        revenues = revenues.filter(date__lte=end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recettes"

    headers = ['Date', 'Catégorie', 'Description', 'Montant (DH)']
    ws.append(headers)

    for revenue in revenues:
        ws.append([
            revenue.date.strftime('%d/%m/%Y'),
            revenue.category,
            revenue.description or '-',
            float(revenue.amount)
        ])

    # Format columns
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=revenus.xlsx'
    wb.save(response)
    return response
